""" This file is to generate lists for the next py file to write the OWL file.
as_assump --> assumption ID list
as_uniq --> combination of gene_id and miRNA name to corespond with the as_list
as_coexpression --> coexpression conclusion list (will be Y or N)

"""
import math


from openpyxl import load_workbook

wb = load_workbook(filename = 'transfering_assumption.xlsx')
ws = wb.get_sheet_by_name(name = 'test')
as_assump = [] # list of assumpID
as_uniq = [] # list of ncbi_ID miRNA_name
varID = 250

for i in range(2,1006):
    target_name = ws.cell('E'+ str(i)).value
    #print target_name
    target_name = str(target_name).strip()
    if target_name <> 'None':
        miRNA_name =  ws.cell('D'+ str(i)).value
        miRNA_name = miRNA_name.strip()
        NCBI_ID = ws.cell('H'+ str(i)).value
        NCBI_ID = str(NCBI_ID)
        uniq = NCBI_ID+ '_' +str(miRNA_name)
        
        #p1 = miRNA_list.miRNA_name_list.index(miRNA_name)
        #miRNA_ID = miRNA_list.miRNA_ID_list[p1]
        if uniq not in as_uniq:
            as_uniq.append( str(uniq) )
            varID = varID+1
            len1 = int(math.log10(varID))+1
            string_val = "0" * (7-len1)
            assump_ID = string_val+str(varID)
            as_assump.append(str(assump_ID))


#p = as_uniq.index('1021_Let-7e')
#print as_assump[p]
#print len(as_uniq)
   

    



















